import bobs
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
import camelot
from itertools import chain

def remove_protections(file):
    wb = load_workbook(file)
    wb.security.lockStructure = False
    for s in wb.sheetnames:
        wb[s].protection.sheet = False
    wb.save(file)

class Michigan(bobs.Table):
    state = 'Michigan'

    def __init__(self, df):
        self.df = df.replace(r'[\*\n]', '', regex=True)
        self.year = re.search(r'\d{4}', self.df.columns[1])[0]

    def clean(self, jump):
        data = []
        # TODO - Totals row should be similar logic.
        for month, *casinos in self.body[:-1].itertuples(index=False):
            idx = 0
            date = datetime.strptime(f'{self.year}-{month}', '%Y-%B')
            while idx < len(casinos) - 2:
                data.append({
                    'State': self.state,
                    'Category': self.category,
                    'Sub-Category': self.subcategory,
                    'Date': date,
                    **self.header.iloc[idx // jump],
                    **self.get_next(casinos, idx)
                })
                idx += jump
        out_df = pd.DataFrame(data)
        return out_df

    def add_single_row(self, data):
        raise NotImplementedError("Specifics of adding a single row should be determined by subclasses.")

class RetailSports(Michigan):
    def __init__(self, link):
        # PDFs are easier to parse than encrypted Excel.
        self.df = self.first_row_to_columns(camelot.read_pdf(link)[0].df).replace('', np.nan)
        self.category = 'Online Sports Betting (OSB)'
        self.subcategory = 'Retail'
        
        super().__init__(self.df)

        self.header = pd.DataFrame(self.df.iloc[0].dropna().str.title()).reset_index(drop=True)
        self.header.columns = ['Provider']
        self.body = (self.first_row_to_columns(
            self.df.iloc[1:16].
            replace(0, np.nan).
            dropna(thresh=4)
        ))
        # Edge case where extra dates are included.
        self.body['Month'] = self.body['Month'].apply(lambda x: x.split(' ')[0])

    def clean(self):
        return super().clean(4)

    def get_next(self, data, idx):
        return {
            'Total Handle': data[idx],
            'Total Gross Receipts': data[idx+1],
            'Adjusted Gross Receipts': data[idx+2],
            'State Tax': data[idx+3]
        }

class InternetSports(Michigan):
    def __init__(self, link):
        self.df = pd.read_excel(link, sheet_name=0)
        self.category = 'Online Sports Betting (OSB)'
        self.subcategory = 'Internet'
        
        super().__init__(self.df)
        
        self.header = (self.df.iloc[:3, 1:-1].
            dropna(how='all', axis=1).
            T.
            reset_index(drop=True)
        )
        self.header.columns = ['Operators', 'Provider', 'Sub-Provider']
        self.body = (self.first_row_to_columns(
            self.df.iloc[4:18].
            replace(0, np.nan).
            dropna(thresh=4)
        ))

    def clean(self):
        return super().clean(4)

    def get_next(self, data, idx):
        return {
            'Total Handle': data[idx],
            'Total Gross Receipts': data[idx+1],
            'Adjusted Gross Receipts': data[idx+2],
            'State Tax': data[idx+3]
        }

class InternetGames(Michigan):
    def __init__(self, link):
        self.df = pd.read_excel(link, sheet_name=0)
        self.category = 'iGaming'
        self.subcategory = None

        super().__init__(self.df)

        self.header = (self.df.iloc[:3, 1:-2].
            dropna(how='all', axis=1).
            T.
            reset_index(drop=True)
        )
        self.header.columns = ['Operators', 'Provider', 'Sub-Provider']
        self.body = (self.first_row_to_columns(
            self.df.iloc[4:18].
            replace(0, np.nan).
            dropna(thresh=4)
        ))

    def clean(self):
        return super().clean(3)

    def get_next(self, data, idx):
        return {
            'Total Gross Receipts': data[idx],
            'Adjusted Gross Receipts': data[idx+1],
            'State Tax': data[idx+2]
        }


if __name__ == '__main__':
    url = 'https://www.michigan.gov/mgcb/detroit-casinos/resources/revenues-and-wagering-tax-information'
    retail_sports = bobs.get_links(url, text_keys=['Retail Sports Betting', 'PDF'])
    internet_sports = bobs.get_links(url, text_keys=['Internet Sports Betting'])
    parsed = chain([RetailSports(r) for r in retail_sports], [InternetSports(i) for i in internet_sports])
    cleaned = [p.clean() for p in parsed]
    df = pd.concat(cleaned)
    df = df[['State', 'Category', 'Sub-Category', 'Date', 'Operators', 'Provider', 'Sub-Provider', 
             'Total Handle', 'Total Gross Receipts', 'Adjusted Gross Receipts', 'State Tax']]
    df.to_excel('Michigan (OSB).xlsx', index=False)

    internet_games = bobs.get_links(url, text_keys=['Internet Gaming', 'Excel'])
    parsed = [
        InternetGames(internet_games[0], 'Internet Gaming 2023'),
        InternetGames(internet_games[1], 'Internet Gaming 2022'),
        InternetGames(internet_games[2], 'Internet Gaming 2021')
    ]
    cleaned = [p.clean() for p in parsed]
    df = pd.concat(cleaned)
    df = df[['State', 'Category', 'Sub-Category', 'Date', 'Operators', 'Provider', 'Sub-Provider', 
            'Total Gross Receipts', 'Adjusted Gross Receipts', 'State Tax']]
    df.to_excel('Michigan (iGaming).xlsx', index=False)
